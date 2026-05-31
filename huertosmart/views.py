# views.py — Vistas principales de HuertoSmart
"""
Este archivo es el CEREBRO de la aplicación Django. Aquí se definen todas las
'vistas', que son funciones Python que reciben una petición del navegador (request)
y devuelven una respuesta (normalmente una página HTML renderizada).

Cada URL del proyecto apunta a una función de este archivo. Cuando el usuario
entra a una página, Django busca qué función debe ejecutar, la ejecuta, y
devuelve el resultado al navegador.

Las vistas de este proyecto NUNCA acceden directamente a la base de datos.
En su lugar, usan los 'repositorios' (carpeta repositories/) que son los
encargados de hablar con la base de datos. Esto se llama Patrón Repository.
"""

# IMPORTS — Herramientas que necesita este archivo

import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.templatetags.static import static

# --- Modelos propios del proyecto ---
from .models import Siembra

# --- Formularios del proyecto ---
from .forms import CultivoFilterForm, HuertoForm, SiembraForm

# --- Repositorios (Patrón Repository) ---
from .repositories import CultivoRepository, HuertoRepository, SiembraRepository, DiagnosticoRepository, EnfermedadRepository

# CONFIGURACIÓN GLOBAL DEL MÓDULO

logger = logging.getLogger(__name__)

# Porcentaje mínimo de confianza que debe tener el modelo de IA para aceptar un diagnóstico
UMBRAL_CONFIANZA_MINIMA = 50

# MAPA DE IMÁGENES DE CULTIVOS

# Diccionario que relaciona el nombre de cada cultivo con la ruta de su imagen
IMAGENES_CULTIVOS = {
    'Tomate':       'img/cultivos/tomate.jpg',
    'Lechuga':      'img/cultivos/lechuga.jpg',
    'Pimiento':     'img/cultivos/pimiento.jpg',
    'Calabacín':    'img/cultivos/calabacin.jpg',
    'Patata':       'img/cultivos/patata.jpg',
    'Zanahoria':    'img/cultivos/zanahoria.jpg',
    'Ajo':          'img/cultivos/ajo.jpg',
    'Cebolla':      'img/cultivos/cebolla.jpg',
    'Espinaca':     'img/cultivos/espinaca.jpg',
    'Acelga':       'img/cultivos/acelga.jpg',
    'Judía verde':  'img/cultivos/judia.jpg',
    'Pepino':       'img/cultivos/pepino.jpg',
    'Berenjena':    'img/cultivos/berenjena.jpg',
    'Rábano':       'img/cultivos/rabano.jpg',
    'Fresa':        'img/cultivos/fresa.jpg',
    'Maíz':         'img/cultivos/maiz.jpg',
    'Manzano':      'img/cultivos/manzano.jpg',
    'Uva':          'img/cultivos/uva.jpg',
    'Melocotonero': 'img/cultivos/melocotonero.jpg',
    'Albahaca':     'img/cultivos/albahaca.jpg',
}

def _imagen_url(nombre):
    """Devuelve la URL completa de la imagen de un cultivo dado su nombre.
    Si el cultivo no tiene imagen asignada, devuelve una cadena vacía."""
    ruta = IMAGENES_CULTIVOS.get(nombre, '')
    return static(ruta) if ruta else ''

# PÁGINAS PÚBLICAS — Accesibles sin iniciar sesión

def home(request):
    """Página de inicio de HuertoSmart."""
    return render(request, 'huertosmart/home.html')

# F3 — BIBLIOTECA DE CULTIVOS
# Funcionalidad pública: cualquier visitante puede consultar los cultivos

def lista_cultivos(request):
    """Lista filtrable de los 20 cultivos del catálogo."""
    form = CultivoFilterForm(request.GET or None)
    repo = CultivoRepository()

    if form.is_valid():
        # El usuario ha aplicado filtros — consultamos solo los cultivos que los cumplen
        cultivos = repo.filtrar(
            busqueda=form.cleaned_data.get('busqueda'),
            dificultad=form.cleaned_data.get('dificultad'),
            exposicion=form.cleaned_data.get('exposicion'),
            riego=form.cleaned_data.get('riego'),
            mes_siembra=form.cleaned_data.get('mes_siembra'),
        )
    else:
        # No hay filtros o el formulario no es válido — mostramos todos los cultivos
        cultivos = repo.get_all()

    # Construimos una lista de diccionarios con el cultivo y su URL de imagen
    cultivos_con_imagen = [
        {'cultivo': c, 'imagen_url': _imagen_url(c.nombre)}
        for c in cultivos
    ]

    return render(request, 'huertosmart/cultivos/lista.html', {
        'form': form,
        'cultivos_con_imagen': cultivos_con_imagen,
        'total': cultivos.count(),  # Número de resultados para mostrar en el template
    })

def detalle_cultivo(request, cultivo_slug):
    """Ficha completa de un cultivo accedida por su slug (URL amigable)."""

    repo = CultivoRepository()
    cultivo = repo.get_by_slug(cultivo_slug)

    # Si el slug no existe en la base de datos, redirigimos a la lista
    if cultivo is None:
        return redirect('huertosmart:lista_cultivos')

    # Los meses se guardan como texto separado por comas: "marzo,abril,mayo"
    # Aquí los convertimos en una lista Python para mostrarlos como etiquetas en el template
    meses_siembra = [m.strip() for m in cultivo.meses_siembra.split(',') if m.strip()]
    meses_cosecha = [m.strip() for m in cultivo.meses_cosecha.split(',') if m.strip()]

    return render(request, 'huertosmart/cultivos/detalle.html', {
        'cultivo': cultivo,
        'imagen_url': _imagen_url(cultivo.nombre),
        'meses_siembra': meses_siembra,
        'meses_cosecha': meses_cosecha,
        'enfermedades': EnfermedadRepository().enfermedades_de_cultivo(cultivo),
    })

# F1 — DIAGNÓSTICO POR FOTO (Funcionalidad de IA — origen curso CE-IABD)
# Requiere login. Integra AWS Rekognition y modelo MobileNetV2 de Hugging Face.

@login_required
def diagnostico_nuevo(request):
    """Vista que gestiona el formulario de diagnóstico de enfermedades por foto."""

    if request.method != 'POST':
        # Si el usuario solo entra a ver la página (GET), mostramos el formulario vacío
        huertos = HuertoRepository().get_huertos_usuario(request.user)
        return render(request, 'huertosmart/diagnostico/nuevo.html', {'huertos': huertos})

    # --- PASO 1: Verificar que se ha subido una imagen ---
    imagen = request.FILES.get('imagen')
    if not imagen:
        messages.error(request, 'Debes seleccionar una imagen.')
        return redirect('huertosmart:diagnostico_nuevo')

    # Leemos los bytes de la imagen para pasarlos a AWS y al modelo de IA
    imagen_bytes = imagen.read()

    # --- PASO 2: Validar con AWS Rekognition que la imagen contiene una planta ---
    es_planta = True
    try:
        from .services.aws_service import get_aws_service
        aws = get_aws_service()
        resultado_rekognition = aws.validar_es_planta(imagen_bytes)
        es_planta = resultado_rekognition['es_planta']
        if not es_planta:
            messages.warning(request, resultado_rekognition['mensaje'])
            return redirect('huertosmart:diagnostico_nuevo')
    except Exception as e:
        # Si AWS falla, registramos el aviso pero dejamos continuar el diagnóstico
        logger.warning(f"Rekognition no disponible, se omite validación: {e}")

    # --- PASO 3: Clasificar la enfermedad con el modelo de IA (Hugging Face) ---
    try:
        from .services.ia_service import get_servicio_ia
        servicio = get_servicio_ia()
        resultado_ia = servicio.diagnosticar(imagen_bytes)
    except Exception as e:
        logger.error(f"Error en modelo IA: {e}")
        messages.error(request, 'El servicio de diagnóstico no está disponible en este momento.')
        return redirect('huertosmart:diagnostico_nuevo')

    nombre_clase = resultado_ia['nombre_clase']
    confianza = resultado_ia['confianza']

    # --- PASO 4: Comprobar que la confianza supera el umbral mínimo ---
    if confianza < UMBRAL_CONFIANZA_MINIMA:
        messages.warning(request, f'La confianza del modelo es muy baja ({confianza}%). Sube una foto más clara.')
        return redirect('huertosmart:diagnostico_nuevo')

    # --- PASO 5: Buscar la enfermedad en la base de datos por el nombre que devuelve el modelo ---
    enfermedad = EnfermedadRepository().buscar_por_nombre_modelo(nombre_clase)

    # --- PASO 6: Vincular el diagnóstico a una siembra si el usuario la seleccionó ---
    siembra_id = request.POST.get('siembra')
    siembra = None
    if siembra_id:
        try:
            siembra = Siembra.objects.get(pk=siembra_id, huerto__usuario=request.user)
        except Siembra.DoesNotExist:
            pass

    # --- PASO 7: Guardar el diagnóstico en la base de datos ---
    imagen.seek(0)

    diagnostico = DiagnosticoRepository().create(
        usuario=request.user,
        siembra=siembra,
        imagen=imagen,
        enfermedad_detectada=enfermedad,
        confianza=confianza,
        es_planta_valida=es_planta,
        notas_usuario=request.POST.get('notas', ''),
    )

    # Redirigimos al detalle del diagnóstico recién creado
    return redirect('huertosmart:diagnostico_detalle', diagnostico_id=diagnostico.pk)

@login_required
def diagnostico_detalle(request, diagnostico_id):
    """Muestra el resultado completo de un diagnóstico."""
    repo = DiagnosticoRepository()
    diagnostico = repo.get_by_id(diagnostico_id)

    # Verificamos que el diagnóstico existe y pertenece al usuario actual
    if diagnostico is None or diagnostico.usuario != request.user:
        messages.error(request, 'No tienes acceso a ese diagnóstico.')
        return redirect('huertosmart:diagnostico_historial')

    return render(request, 'huertosmart/diagnostico/detalle.html', {
        'diagnostico': diagnostico,
        # Si la confianza es menor del 70%, el template muestra un aviso visible
        'confianza_baja': diagnostico.confianza < 70,
    })

@login_required
def diagnostico_historial(request):
    """Lista todos los diagnósticos realizados por el usuario."""
    repo = DiagnosticoRepository()
    return render(request, 'huertosmart/diagnostico/historial.html', {
        'diagnosticos': repo.get_diagnosticos_usuario(request.user),
    })

# F4 — PANEL DEL HUERTO
# Funcionalidad privada: cada usuario solo ve y gestiona sus propios huertos

@login_required
def mi_huerto(request):
    """Panel principal del usuario. Lista sus huertos o redirige a crear el primero."""
    repo = HuertoRepository()
    huertos = repo.get_huertos_usuario(request.user)

    if not huertos.exists():
        # Si el usuario aún no tiene huertos, lo llevamos directamente a crear uno
        messages.info(request, 'Todavía no tienes ningún huerto. ¡Crea el primero!')
        return redirect('huertosmart:crear_huerto')

    return render(request, 'huertosmart/huerto/panel.html', {'huertos': huertos})

@login_required
def crear_huerto(request):
    """Formulario para crear un nuevo huerto."""
    if request.method == 'POST':
        form = HuertoForm(request.POST)
        if form.is_valid():
            huerto = form.save(commit=False)

            huerto.usuario = request.user  # Asignamos el propietario del huerto
            huerto.save()
            messages.success(request, f'Huerto "{huerto.nombre}" creado correctamente.')
            return redirect('huertosmart:detalle_huerto', huerto_slug=huerto.slug)
    else:
        form = HuertoForm()

    return render(request, 'huertosmart/huerto/crear.html', {'form': form})

@login_required
def detalle_huerto(request, huerto_slug):
    """Detalle completo de un huerto: siembras, filtros y previsión meteorológica.

    LÓGICA DE NEGOCIO: solo el propietario puede ver su huerto.
    """
    siembra_repo = SiembraRepository()

    # Buscamos el huerto por slug — get_by_slug solo devuelve huertos activos
    huerto = HuertoRepository().get_by_slug(huerto_slug)
    if huerto is None or huerto.usuario != request.user:
        messages.error(request, 'No tienes acceso a ese huerto.')
        return redirect('huertosmart:mi_huerto')

    # --- Filtro de siembras por estado ---
    estado_filtro = request.GET.get('estado', '')
    siembras = (
        siembra_repo.filtrar_por_estado(huerto, estado_filtro)
        if estado_filtro
        else siembra_repo.get_siembras_huerto(huerto)
    )

    # --- Previsión meteorológica con AEMET (Funcionalidad CE-IABD) ---
    prediccion, alertas = [], []
    if huerto.codigo_postal:
        try:
            from .services.aemet_service import get_alertas_huerto
            resultado = get_alertas_huerto(huerto.codigo_postal)
            prediccion = resultado['prediccion']
            alertas = resultado['alertas']
        except Exception as e:
            # Si AEMET falla, la página sigue funcionando sin la previsión
            logger.warning(f"AEMET no disponible: {e}")

    return render(request, 'huertosmart/huerto/detalle.html', {
        'huerto': huerto,
        'siembras': siembras,
        'estado_filtro': estado_filtro,
        # ESTADO_CHOICES es la lista de estados posibles definida en el modelo Siembra
        'estados': [('', 'Todas')] + list(Siembra.ESTADO_CHOICES),
        'prediccion': prediccion,
        'alertas': alertas,
    })

@login_required
def eliminar_huerto(request, huerto_slug):
    """Elimina un huerto del usuario previa confirmación en pantalla.

    LÓGICA DE NEGOCIO: solo el propietario puede eliminar su huerto.
    Solo acepta POST para evitar eliminaciones accidentales por GET.
    """
    huerto = HuertoRepository().get_by_slug(huerto_slug)
    if huerto is None or huerto.usuario != request.user:
        messages.error(request, 'No tienes acceso a ese huerto.')
        return redirect('huertosmart:mi_huerto')

    if request.method == 'POST':
        # El usuario ha confirmado la eliminación en el formulario de confirmación
        nombre = huerto.nombre
        huerto.delete()
        messages.success(request, f'El huerto "{nombre}" ha sido eliminado.')
        return redirect('huertosmart:mi_huerto')

    return render(request, 'huertosmart/huerto/confirmar_eliminar.html', {'huerto': huerto})

@login_required
def nueva_siembra(request, huerto_slug):
    """Formulario para registrar una nueva siembra en un huerto."""
    huerto = HuertoRepository().get_by_slug(huerto_slug)

    if huerto is None or huerto.usuario != request.user:
        messages.error(request, 'No tienes acceso a ese huerto.')
        return redirect('huertosmart:mi_huerto')

    if request.method == 'POST':
        form = SiembraForm(request.POST)
        if form.is_valid():
            siembra = form.save(commit=False)
            siembra.huerto = huerto  # Vinculamos la siembra al huerto actual
            siembra.save()
            messages.success(request, f'Siembra de {siembra.cultivo.nombre} registrada.')
            return redirect('huertosmart:detalle_huerto', huerto_slug=huerto.slug)
    else:
        form = SiembraForm()

    return render(request, 'huertosmart/huerto/nueva_siembra.html', {'form': form, 'huerto': huerto})

# EXPORTAR A EXCEL — Requisito Django nº 9
# Genera un archivo .xlsx con una hoja por cada huerto del usuario

@login_required
def exportar_excel(request):
    """Genera y descarga un archivo Excel con todas las siembras del usuario."""
    siembra_repo = SiembraRepository()
    huerto_repo = HuertoRepository()
    huertos = huerto_repo.get_huertos_usuario(request.user)

    # Creamos un libro Excel nuevo en memoria
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminamos la hoja vacía que crea openpyxl por defecto

    # --- Estilos para las celdas de cabecera ---
    estilo_cabecera = Font(bold=True, color='FFFFFF')           # Texto blanco y negrita
    fondo_cabecera = PatternFill(fill_type='solid', fgColor='1F6B3A')  # Verde oscuro
    fondo_par = PatternFill(fill_type='solid', fgColor='E8F5E9')       # Verde muy suave para filas pares
    alineacion_centro = Alignment(horizontal='center')

    # Definición de columnas: (nombre de la columna, ancho en caracteres)
    columnas = [
        ('Cultivo', 20), ('Fecha siembra', 15), ('Cosecha estimada', 18),
        ('Cosecha real', 15), ('Cantidad (plantas)', 18), ('Estado', 15), ('Notas', 40),
    ]

    # --- Creamos una hoja por cada huerto ---
    for huerto in huertos:
        ws = wb.create_sheet(title=huerto.nombre[:31])

        # Cabecera con nombres de columnas y estilos
        for col_idx, (nombre_col, ancho) in enumerate(columnas, start=1):
            celda = ws.cell(row=1, column=col_idx, value=nombre_col)
            celda.font = estilo_cabecera
            celda.fill = fondo_cabecera
            celda.alignment = alineacion_centro
            ws.column_dimensions[celda.column_letter].width = ancho

        # Filas de datos: una por cada siembra del huerto
        siembras = siembra_repo.get_siembras_huerto(huerto)
        for fila_idx, siembra in enumerate(siembras, start=2):
            valores = [
                siembra.cultivo.nombre,
                siembra.fecha_siembra.strftime('%d/%m/%Y') if siembra.fecha_siembra else '',
                siembra.fecha_cosecha_estimada.strftime('%d/%m/%Y') if siembra.fecha_cosecha_estimada else '',
                siembra.fecha_cosecha_real.strftime('%d/%m/%Y') if siembra.fecha_cosecha_real else '',
                siembra.cantidad,
                siembra.get_estado_display(),  # Devuelve el texto legible del estado (ej: "En crecimiento")
                siembra.notas,
            ]

            for col_idx, valor in enumerate(valores, start=1):
                celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
                # Aplicamos fondo verde suave a las filas pares para mejor legibilidad
                if fila_idx % 2 == 0:
                    celda.fill = fondo_par

        # Si el huerto no tiene siembras, añadimos una fila informativa
        if not siembras.exists():
            ws.cell(row=2, column=1, value='Este huerto no tiene siembras registradas.')

    # Si el usuario no tiene ningún huerto, creamos una hoja con mensaje informativo
    if not huertos.exists():
        ws = wb.create_sheet(title='Sin datos')
        ws.cell(row=1, column=1, value='No tienes huertos registrados en HuertoSmart.')

    # --- Convertimos el libro Excel a bytes y lo enviamos al navegador ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="huertosmart_{request.user.username}.xlsx"'
    return response
